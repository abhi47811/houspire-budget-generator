import math, openpyxl
from openpyxl.styles import Font
from config import VENDOR_COL_WIDTHS

BOLD  = Font(name="Calibri", size=12, bold=True)
PLAIN = Font(name="Calibri", size=12, bold=False)


def _km(lat, lng, lat0, lng0):
    dlat = (lat - lat0) * 111.0
    dlng = (lng - lng0) * 111.0 * math.cos(math.radians(lat0))
    return round(math.hypot(dlat, dlng), 1)


def write_vendor_excel(project_name, city, pincode, centroid, vendors, notes_text, output_path):
    lat0, lng0 = centroid
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vendors"

    headers = ["Category", "Vendor", "Specialty / Brands", "Area",
               f"Approx. km from {pincode}", "Rating (count)", "Phone"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h).font = BOLD

    ws.freeze_panes = "A2"
    for col, w in VENDOR_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    r = 2
    for v in vendors:
        dist = _km(v.lat, v.lng, lat0, lng0) if v.lat and v.lng else ""
        for c, val in enumerate(
            [v.category, v.vendor, v.specialty, v.area, dist, v.rating, v.phone], 1
        ):
            ws.cell(r, c, val).font = PLAIN
        r += 1

    ns = wb.create_sheet("Notes")
    ns.column_dimensions["A"].width = 120
    for i, line in enumerate(notes_text.splitlines(), 1):
        ns.cell(i, 1, line)

    wb.save(output_path)
    return output_path
