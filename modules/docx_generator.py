import os, json, subprocess, tempfile
import openpyxl

BRAND_PRIMARY = "1B4D3E"
BRAND_ACCENT  = "D4AF37"
BRAND_LIGHT   = "F5F0E8"
ROW_ALT       = "F7F5F0"


def _read_boq_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["BOQ Template"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None and r[1] is None:
            continue
        rows.append({
            "category": r[0] or "",
            "description": r[1] or "",
            "unit": r[2] or "",
            "qty": r[3] or 0,
            "rate": r[4] or 0,
        })
    ws2 = wb["Rate Sources"]
    sources = []
    for r in ws2.iter_rows(min_row=2, values_only=True):
        if r[0]:
            sources.append({"item": r[0] or "", "basis": r[1] or "", "source": r[2] or ""})
    return rows, sources


def _read_vendor_excel(excel_path):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Vendors"]
    vendors = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        vendors.append({
            "category": r[0] or "",
            "vendor": r[1] or "",
            "specialty": r[2] or "",
            "area": r[3] or "",
            "dist": r[4] or "",
            "rating": r[5] or "",
            "phone": r[6] or "",
        })
    notes = ""
    if "Notes" in wb.sheetnames:
        ns = wb["Notes"]
        notes = "\n".join(str(r[0].value or "") for r in ns.iter_rows())
    return vendors, notes


def _node_script_boq(data):
    return f"""
const {{ Document, Packer, Paragraph, TextRun, Table, TableRow, TableCell,
        Header, Footer, AlignmentType, BorderStyle, WidthType, ShadingType,
        PageBreak, PageNumber }} = require("docx");
const fs = require("fs");

const data = {json.dumps(data, ensure_ascii=False)};

const PRIMARY = "1B4D3E";
const ACCENT  = "D4AF37";
const LIGHT   = "F5F0E8";
const ALT     = "F7F5F0";

function headerPara(text) {{
  return new Paragraph({{
    children: [new TextRun({{ text, font: "Arial", size: 20, color: "888888", italics: true }})],
    alignment: AlignmentType.RIGHT,
  }});
}}

function footerPara(pageNum) {{
  return new Paragraph({{
    children: [
      new TextRun({{ text: "www.houspire.ai  |  Page ", font: "Arial", size: 18 }}),
      new TextRun({{ children: [PageNumber.CURRENT], font: "Arial", size: 18 }}),
    ],
    alignment: AlignmentType.CENTER,
  }});
}}

function coverPage(client, city, tier, rowCount) {{
  return [
    new Paragraph({{ spacing: {{ before: 2400 }} }}),
    new Paragraph({{
      children: [new TextRun({{ text: "HOUSPIRE", font: "Arial", size: 176, bold: true, color: PRIMARY, characterSpacing: 300 }})],
      alignment: AlignmentType.CENTER,
    }}),
    new Paragraph({{
      children: [new TextRun({{ text: "Design. Deliver. Delight.", font: "Arial", size: 28, color: ACCENT, italics: true }})],
      alignment: AlignmentType.CENTER,
    }}),
    new Paragraph({{
      border: {{ bottom: {{ color: ACCENT, size: 6, style: BorderStyle.SINGLE }} }},
      spacing: {{ after: 200 }},
    }}),
    new Paragraph({{ spacing: {{ before: 400 }} }}),
    new Paragraph({{
      children: [new TextRun({{ text: "BILL OF QUANTITIES", font: "Arial", size: 52, bold: true, color: PRIMARY }})],
      alignment: AlignmentType.CENTER,
    }}),
    new Paragraph({{
      children: [new TextRun({{ text: `${{client}} | ${{city}} | ${{tier}}`, font: "Arial", size: 28, color: "555555" }})],
      alignment: AlignmentType.CENTER,
      spacing: {{ before: 200 }},
    }}),
    new Paragraph({{
      children: [new TextRun({{ text: `${{rowCount}} line items  •  2025-26 rates`, font: "Arial", size: 22, color: "888888" }})],
      alignment: AlignmentType.CENTER,
      spacing: {{ before: 200 }},
    }}),
    new Paragraph({{ children: [new PageBreak()] }}),
  ];
}}

function boqTable(rows) {{
  const headers = ["Category","Description","Unit","Qty","Rate (₹)","Amount (₹)"];
  const widths  = [1800, 4200, 900, 800, 1100, 1200];
  const headerRow = new TableRow({{
    children: headers.map((h, i) => new TableCell({{
      width: {{ size: widths[i], type: WidthType.DXA }},
      shading: {{ fill: PRIMARY, type: ShadingType.CLEAR }},
      children: [new Paragraph({{ children: [new TextRun({{ text: h, font: "Arial", size: 20, bold: true, color: "FFFFFF" }})] }})],
    }})),
    tableHeader: true,
  }});
  const dataRows = rows.map((r, idx) => {{
    const amt = (r.qty && r.rate) ? `₹${{(r.qty * r.rate).toLocaleString("en-IN")}}` : "";
    const vals = [r.category, r.description, r.unit, String(r.qty||""), r.rate ? `₹${{r.rate.toLocaleString("en-IN")}}` : "", amt];
    const fill = idx % 2 === 1 ? ALT : "FFFFFF";
    return new TableRow({{
      children: vals.map((v, i) => new TableCell({{
        width: {{ size: widths[i], type: WidthType.DXA }},
        shading: {{ fill, type: ShadingType.CLEAR }},
        children: [new Paragraph({{ children: [new TextRun({{ text: v, font: "Arial", size: 18 }})] }})],
      }})),
    }});
  }});
  return new Table({{ rows: [headerRow, ...dataRows], width: {{ size: 100, type: WidthType.PERCENTAGE }} }});
}}

function sourcesTable(sources) {{
  const headers = ["Category / Item", "Rate Basis", "Source"];
  const widths  = [2500, 3000, 4500];
  const headerRow = new TableRow({{
    children: headers.map((h, i) => new TableCell({{
      width: {{ size: widths[i], type: WidthType.DXA }},
      shading: {{ fill: PRIMARY, type: ShadingType.CLEAR }},
      children: [new Paragraph({{ children: [new TextRun({{ text: h, font: "Arial", size: 20, bold: true, color: "FFFFFF" }})] }})],
    }})),
    tableHeader: true,
  }});
  const dataRows = sources.map((s, idx) => {{
    const fill = idx % 2 === 1 ? ALT : "FFFFFF";
    return new TableRow({{
      children: [s.item, s.basis, s.source].map((v, i) => new TableCell({{
        width: {{ size: widths[i], type: WidthType.DXA }},
        shading: {{ fill, type: ShadingType.CLEAR }},
        children: [new Paragraph({{ children: [new TextRun({{ text: v||"", font: "Arial", size: 16 }})] }})],
      }})),
    }});
  }});
  return new Table({{ rows: [headerRow, ...dataRows], width: {{ size: 100, type: WidthType.PERCENTAGE }} }});
}}

function disclaimerPage() {{
  return [
    new Paragraph({{ children: [new PageBreak()] }}),
    new Paragraph({{
      children: [new TextRun({{ text: "Disclaimer", font: "Arial", size: 36, bold: true, color: PRIMARY }})],
      spacing: {{ before: 400, after: 200 }},
    }}),
    new Paragraph({{
      children: [new TextRun({{ text: "Houspire operates on a flat-fee model. We do not earn commissions from any vendor, supplier, or brand mentioned in this document. All rates are indicative, researched for the local market (2025-26), and must be confirmed with vendors before ordering. Room sizes are estimated from design renders and should be verified on-site before procurement.", font: "Arial", size: 20, color: "555555" }})],
    }}),
  ];
}}

async function main() {{
  const sections = [];
  const bodyChildren = [
    ...coverPage(data.client, data.city, data.tier, data.rows.length),
    new Paragraph({{ children: [new TextRun({{ text: "Bill of Quantities", font: "Arial", size: 36, bold: true, color: PRIMARY }})], spacing: {{ before: 400, after: 300 }} }}),
    boqTable(data.rows),
    new Paragraph({{ spacing: {{ before: 400, after: 200 }} }}),
    new Paragraph({{ children: [new TextRun({{ text: "Rate Sources", font: "Arial", size: 32, bold: true, color: PRIMARY }})], spacing: {{ before: 400, after: 300 }} }}),
    sourcesTable(data.sources),
    ...disclaimerPage(),
  ];

  const doc = new Document({{
    sections: [{{
      properties: {{
        page: {{
          size: {{ width: 11906, height: 16838 }},
          margin: {{ top: 1000, right: 1000, bottom: 1000, left: 1000 }},
        }},
      }},
      headers: {{
        default: new Header({{ children: [headerPara(`Houspire BOQ | ${{data.client}} | Confidential`)] }}),
      }},
      footers: {{
        default: new Footer({{ children: [footerPara()] }}),
      }},
      children: bodyChildren,
    }}],
  }});

  const buffer = await Packer.toBuffer(doc);
  fs.writeFileSync(data.output_path, buffer);
  console.log("OK:" + data.output_path);
}}
main().catch(e => {{ console.error(e); process.exit(1); }});
"""


def _node_script_vendor(data):
    return f"""
const {{ Document, Packer, Paragraph, TextRun, Table, TableRow, TableCell,
        Header, Footer, AlignmentType, BorderStyle, WidthType, ShadingType,
        PageBreak, PageNumber }} = require("docx");
const fs = require("fs");

const data = {json.dumps(data, ensure_ascii=False)};

const PRIMARY = "1B4D3E";
const ACCENT  = "D4AF37";
const ALT     = "F7F5F0";

function headerPara(text) {{
  return new Paragraph({{
    children: [new TextRun({{ text, font: "Arial", size: 20, color: "888888", italics: true }})],
    alignment: AlignmentType.RIGHT,
  }});
}}

function footerPara() {{
  return new Paragraph({{
    children: [
      new TextRun({{ text: "www.houspire.ai  |  Page ", font: "Arial", size: 18 }}),
      new TextRun({{ children: [PageNumber.CURRENT], font: "Arial", size: 18 }}),
    ],
    alignment: AlignmentType.CENTER,
  }});
}}

function coverPage(client, city, count) {{
  return [
    new Paragraph({{ spacing: {{ before: 2400 }} }}),
    new Paragraph({{
      children: [new TextRun({{ text: "HOUSPIRE", font: "Arial", size: 176, bold: true, color: PRIMARY, characterSpacing: 300 }})],
      alignment: AlignmentType.CENTER,
    }}),
    new Paragraph({{
      children: [new TextRun({{ text: "Design. Deliver. Delight.", font: "Arial", size: 28, color: ACCENT, italics: true }})],
      alignment: AlignmentType.CENTER,
    }}),
    new Paragraph({{
      border: {{ bottom: {{ color: ACCENT, size: 6, style: BorderStyle.SINGLE }} }},
      spacing: {{ after: 200 }},
    }}),
    new Paragraph({{ spacing: {{ before: 400 }} }}),
    new Paragraph({{
      children: [new TextRun({{ text: "LOCAL VENDOR DIRECTORY", font: "Arial", size: 52, bold: true, color: PRIMARY }})],
      alignment: AlignmentType.CENTER,
    }}),
    new Paragraph({{
      children: [new TextRun({{ text: `${{client}} | ${{city}}`, font: "Arial", size: 28, color: "555555" }})],
      alignment: AlignmentType.CENTER,
      spacing: {{ before: 200 }},
    }}),
    new Paragraph({{
      children: [new TextRun({{ text: `${{count}} verified vendors`, font: "Arial", size: 22, color: "888888" }})],
      alignment: AlignmentType.CENTER,
      spacing: {{ before: 200 }},
    }}),
    new Paragraph({{ children: [new PageBreak()] }}),
  ];
}}

function vendorTable(vendors) {{
  const headers = ["Category","Vendor","Specialty / Brands","Area","km","Rating","Phone"];
  const widths  = [1500, 2000, 3000, 2000, 600, 1000, 1400];
  const headerRow = new TableRow({{
    children: headers.map((h, i) => new TableCell({{
      width: {{ size: widths[i], type: WidthType.DXA }},
      shading: {{ fill: PRIMARY, type: ShadingType.CLEAR }},
      children: [new Paragraph({{ children: [new TextRun({{ text: h, font: "Arial", size: 18, bold: true, color: "FFFFFF" }})] }})],
    }})),
    tableHeader: true,
  }});
  const dataRows = vendors.map((v, idx) => {{
    const fill = idx % 2 === 1 ? ALT : "FFFFFF";
    const vals = [v.category, v.vendor, v.specialty, v.area, String(v.dist||""), v.rating, v.phone];
    return new TableRow({{
      children: vals.map((val, i) => new TableCell({{
        width: {{ size: widths[i], type: WidthType.DXA }},
        shading: {{ fill, type: ShadingType.CLEAR }},
        children: [new Paragraph({{ children: [new TextRun({{ text: val||"", font: "Arial", size: 17 }})] }})],
      }})),
    }});
  }});
  return new Table({{ rows: [headerRow, ...dataRows], width: {{ size: 100, type: WidthType.PERCENTAGE }} }});
}}

async function main() {{
  const bodyChildren = [
    ...coverPage(data.client, data.city, data.vendors.length),
    new Paragraph({{ children: [new TextRun({{ text: "Verified Local Vendors", font: "Arial", size: 36, bold: true, color: PRIMARY }})], spacing: {{ before: 400, after: 300 }} }}),
    vendorTable(data.vendors),
  ];

  const doc = new Document({{
    sections: [{{
      properties: {{
        page: {{
          size: {{ width: 11906, height: 16838 }},
          margin: {{ top: 1200, right: 1200, bottom: 1200, left: 1200 }},
        }},
      }},
      headers: {{
        default: new Header({{ children: [headerPara(`Houspire Vendors | ${{data.client}} | Confidential`)] }}),
      }},
      footers: {{
        default: new Footer({{ children: [footerPara()] }}),
      }},
      children: bodyChildren,
    }}],
  }});

  const buffer = await Packer.toBuffer(doc);
  fs.writeFileSync(data.output_path, buffer);
  console.log("OK:" + data.output_path);
}}
main().catch(e => {{ console.error(e); process.exit(1); }});
"""


_PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _run_node_script(script: str) -> None:
    """Write script to project dir and run it so local/global node_modules resolve."""
    script_path = os.path.join(_PROJECT_DIR, "_docx_tmp.js")
    try:
        with open(script_path, "w") as f:
            f.write(script)
        env = os.environ.copy()
        # Ensure global node_modules on NODE_PATH as fallback
        node_global = subprocess.run(
            ["npm", "root", "-g"], capture_output=True, text=True
        ).stdout.strip()
        existing = env.get("NODE_PATH", "")
        env["NODE_PATH"] = f"{node_global}:{existing}" if existing else node_global
        result = subprocess.run(
            ["node", script_path], capture_output=True, text=True,
            cwd=_PROJECT_DIR, env=env
        )
        if result.returncode != 0:
            raise RuntimeError(f"Node.js failed:\n{result.stderr}")
    finally:
        if os.path.exists(script_path):
            os.unlink(script_path)


def generate_branded_boq_docx(client_name, city, tier, excel_path, output_path):
    rows, sources = _read_boq_excel(excel_path)
    data = {
        "client": client_name, "city": city, "tier": tier,
        "rows": rows, "sources": sources, "output_path": output_path,
    }
    _run_node_script(_node_script_boq(data))
    return output_path


def generate_branded_vendor_docx(client_name, city, vendors, output_path):
    vendor_list = [
        {"category": v.category, "vendor": v.vendor, "specialty": v.specialty,
         "area": v.area, "dist": "", "rating": v.rating, "phone": v.phone}
        for v in vendors
    ]
    data = {
        "client": client_name, "city": city,
        "vendors": vendor_list, "output_path": output_path,
    }
    _run_node_script(_node_script_vendor(data))
    return output_path
