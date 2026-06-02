import openpyxl
from openpyxl.styles import Font
from config import BOQ_COL_WIDTHS

BOLD  = Font(name="Calibri", size=12, bold=True)
PLAIN = Font(name="Calibri", size=12, bold=False)

BOQ_HEADERS = ["Category", "Description", "Unit", "Quantity", "Rate", "Amount (Auto-calculated)"]
SRC_HEADERS = ["Category / Item", "Rate Basis", "Source"]


def write_budget_excel(client_name, city, pincode, tier, rows, sources, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheet 1: BOQ Template
    ws = wb.create_sheet("BOQ Template")
    for c, h in enumerate(BOQ_HEADERS, 1):
        ws.cell(1, c, h).font = BOLD
    for i, row in enumerate(rows, 2):
        ws.cell(i, 1, row.category).font = PLAIN
        ws.cell(i, 2, row.description).font = PLAIN
        ws.cell(i, 3, row.unit).font = PLAIN
        ws.cell(i, 4, row.qty).font = PLAIN
        ws.cell(i, 5, row.rate).font = PLAIN
        ws.cell(i, 6, f"=D{i}*E{i}").font = PLAIN
    for col, w in BOQ_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w

    # Sheet 2: Rate Sources
    ws2 = wb.create_sheet("Rate Sources")
    for c, h in enumerate(SRC_HEADERS, 1):
        ws2.cell(1, c, h).font = BOLD
    for i, src in enumerate(sources, 2):
        ws2.cell(i, 1, src.item).font = PLAIN
        ws2.cell(i, 2, src.basis).font = PLAIN
        ws2.cell(i, 3, src.source).font = PLAIN
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 45

    wb.save(output_path)
    return output_path
