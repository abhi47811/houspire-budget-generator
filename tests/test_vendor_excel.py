import tempfile, os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from modules.vendor_excel import write_vendor_excel
from modules.vendor_finder import VendorRow


def _make_vendors():
    return [
        VendorRow("Flooring", "Kajaria World", "Large-format vitrified", "Banjara Hills, Hyderabad (500034)", 17.415, 78.448, "4.7 (210)", "+91 99999 11111"),
        VendorRow("Flooring", "Somany Tiles", "Polished porcelain", "Jubilee Hills, Hyderabad (500033)", 17.430, 78.440, "4.5 (88)", "NA - visit showroom"),
        VendorRow("Hardware", "Hettich Dealer", "Hinges, slides, lifts", "Secunderabad (500003)", 17.443, 78.499, "4.6 (55)", "+91 88888 22222"),
    ]


def test_distance_is_float():
    vendors = _make_vendors()
    centroid = (17.420, 78.450)
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test_vendor.xlsx")
        write_vendor_excel("Test Project", "Hyderabad", "500032", centroid, vendors, "Notes here", path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Vendors"]
        for r in range(2, 2 + len(vendors)):
            dist_val = ws.cell(r, 5).value
            assert isinstance(dist_val, float), f"Row {r}: distance should be float, got {type(dist_val)}"


def test_header_bold():
    vendors = _make_vendors()
    centroid = (17.420, 78.450)
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test_vendor.xlsx")
        write_vendor_excel("Test Project", "Hyderabad", "500032", centroid, vendors, "Notes here", path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Vendors"]
        assert ws.cell(1, 1).font.bold is True
        assert ws.cell(2, 1).font.bold is False


def test_notes_sheet_present():
    vendors = _make_vendors()
    centroid = (17.420, 78.450)
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test_vendor.xlsx")
        write_vendor_excel("Test Project", "Hyderabad", "500032", centroid, vendors, "Line1\nLine2", path)
        wb = openpyxl.load_workbook(path)
        assert "Notes" in wb.sheetnames


def test_row_count_matches_vendors():
    vendors = _make_vendors()
    centroid = (17.420, 78.450)
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test_vendor.xlsx")
        write_vendor_excel("Test Project", "Hyderabad", "500032", centroid, vendors, "Notes", path)
        wb = openpyxl.load_workbook(path)
        ws = wb["Vendors"]
        assert ws.max_row == 1 + len(vendors)
