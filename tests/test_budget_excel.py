import tempfile, os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))

import openpyxl
from modules.budget_excel import write_budget_excel
from modules.boq_generator import BOQRow, RateSource


def _make_rows():
    return [
        BOQRow("Flooring", "Premium vitrified tile Kajaria - Living Room", "sft", 200.0, 179.0),
        BOQRow("Electrical", "6A switch Legrand Arteor - Living Room", "nos", 4.0, 1150.0),
        BOQRow("HVAC", "Split AC 1.5T Daikin - Master Bedroom", "nos", 1.0, 51000.0),
        BOQRow("HVAC", "Split AC 1.5T install kit - Master Bedroom", "lump", 1.0, 11000.0),
    ]


def _make_sources():
    return [
        RateSource("Flooring", "Kajaria 2025-26 Hyderabad", "kajariaceramics.com"),
        RateSource("Note", "Hyderabad multiplier applied silently.", "-"),
    ]


def test_formula_not_hardcoded():
    rows = _make_rows()
    sources = _make_sources()
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test_boq.xlsx")
        write_budget_excel("Test Client", "Hyderabad", "500032", "Mid-tier", rows, sources, path)
        wb = openpyxl.load_workbook(path)  # NOT data_only — preserves formulas
        ws = wb["BOQ Template"]
        for r in range(2, 2 + len(rows)):
            val = ws.cell(r, 6).value
            assert val == f"=D{r}*E{r}", f"Row {r}: expected formula =D{r}*E{r}, got {val!r}"


def test_no_footer_rows():
    rows = _make_rows()
    sources = _make_sources()
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test_boq.xlsx")
        write_budget_excel("Test Client", "Hyderabad", "500032", "Mid-tier", rows, sources, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["BOQ Template"]
        assert ws.max_row == 1 + len(rows), f"Expected {1 + len(rows)} rows, got {ws.max_row}"


def test_header_bold_data_not_bold():
    rows = _make_rows()
    sources = _make_sources()
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test_boq.xlsx")
        write_budget_excel("Test Client", "Hyderabad", "500032", "Mid-tier", rows, sources, path)
        wb = openpyxl.load_workbook(path)
        ws = wb["BOQ Template"]
        assert ws.cell(1, 1).font.bold is True
        assert ws.cell(2, 1).font.bold is False


def test_two_sheets():
    rows = _make_rows()
    sources = _make_sources()
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "test_boq.xlsx")
        write_budget_excel("Test Client", "Hyderabad", "500032", "Mid-tier", rows, sources, path)
        wb = openpyxl.load_workbook(path)
        assert "BOQ Template" in wb.sheetnames
        assert "Rate Sources" in wb.sheetnames
